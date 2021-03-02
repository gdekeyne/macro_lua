
"""
The Macro comparing the LUA with the autocad file
"""

from copy import copy

import numpy as np
import pandas as pd

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class Macro:
    def __init__(self, lua, autocad, box_name):
        # INITIALIZING INPUT AND OUTPUT
        # Reading the input files
        self.lua = pd.read_excel(lua, skiprows=1, sheet_name='IO')
        self.autocad = pd.read_excel(autocad)
        # Creating the result sheet
        self.results = openpyxl.load_workbook(autocad)
        self.sheet = self.results.worksheets[0]
        # Arranging input data
        self.arrange_autocad()
        self.arrange_lua()
        self.n_col = len(self.autocad.columns)
        self.last_row = self.find_last_row()
        # Sorting the columns in the result sheet
        self.arrange_sheet()

        # MISCELLANEOUS
        self.box_name = box_name.lower().replace(' ', '_')
        if self.box_name not in self.lua.keys():
            raise ValueError('{} not found in LUA box names'.format(self.box_name))

        # Sheets starts counts at 1, 1st line ignored, 2nd line is column names
        self.index_offset = 3
        self.green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        self.yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        self.orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    def arrange_sheet(self):
        for i in range(6):
            self.sheet.insert_cols(self.n_col)

        for n, key in enumerate(['index', 'dutch', 'french', 'signal', 'polarity']):
            try:
                list(self.sheet.columns)[n + self.n_col][0].value = key
            except IndexError:
                raise IndexError('Index {} beyond length of autocad file which is {}'
                                 .format(n + self.n_col, len(list(self.sheet.columns))))

        # Inserting columns at the end of the file inserts it at the second to last place
        # Hence the configuration is at the last column and needs to be moved back
        for cell_config, cell_last_column in zip(list(self.sheet.columns)[self.n_col-1], list(self.sheet.columns)[-1]):
            cell_config.value = cell_last_column.value
            self.copy_style(cell_config, cell_last_column)
            cell_last_column.value = ''

    @staticmethod
    def copy_style(cell_config, cell_last_column):
        if cell_last_column.has_style:
            cell_config.font, cell_last_column.font = copy(cell_last_column.font), copy(cell_config.font)
            cell_config.border, cell_last_column.border = copy(cell_last_column.border), copy(cell_config.border)
            cell_config.fill, cell_last_column.fill = copy(cell_last_column.fill), copy(cell_config.fill)
            cell_config.number_format, cell_last_column.number_format = copy(cell_last_column.number_format), copy(cell_config.number_format)
            cell_config.protection, cell_last_column.protection = copy(cell_last_column.protection), copy(cell_config.protection)
            cell_config.alignment, cell_last_column.alignment = copy(cell_last_column.alignment), copy(cell_config.alignment)

    def arrange_autocad(self):
        for key in self.autocad.columns:
            self.autocad.rename(columns={key: key.lower().replace(' ', '_')}, inplace=True)

        for key in ['i/o_descr_nl', 'i/o_descr_fr', 'i/o_conn_02_function', 'i/o_conn_01_function',
                    'eq_name', 'i/o_name', 'standard_configuration_id']:
            if key not in self.autocad.columns:
                raise KeyError('Key {} not found in autocad'.format(key))

    def arrange_lua(self):
        for key in self.lua.columns:
            self.lua.rename(columns={key: key.lower().replace(' ', '_')}, inplace=True)

        for key in ['i/o_name', 'i/o_type', 'interface_device_name', 'standard_configuration_id',
                    'infolist/princ.diagr._name_nl', 'infolist/princ.diagr._name_fr', 'signal_code', 'polarity']:
            if key not in self.lua.columns:
                raise KeyError('Key {} not found in LUA reference'.format(key))

        # Reassembling the name from the type and version number
        full_name_list = list()
        for io_name, io_type in zip(self.lua['i/o_name'], self.lua['i/o_type']):
            if (io_type == 'BO') or (io_type == 'BI'):
                if str(io_name).replace('.', '').isdigit():
                    full_name_list.append(io_type[-1] + str(io_name))
                else:
                    full_name_list.append(io_name)
            else:
                full_name_list.append(np.nan)

        self.lua['full_name'] = full_name_list

    def find_last_row(self):
        """
        Finds the first row whose cells are all empty
        """
        last_row = 0
        for row in self.autocad.iterrows():
            if all([str(cell) == 'nan' for cell in row[1]]):
                break
            last_row = row[0]

        return last_row

    def get_indexes(self):
        """
        Finds the list of the cells in the LUA matching each element of the autocad file
        """
        for row, (eq_name, io_name, conf_id) in enumerate(zip(self.autocad['eq_name'],
                                                              self.autocad['i/o_name'],
                                                              self.autocad['standard_configuration_id'])):
            if row > self.last_row:
                break

            index_eq = self.get_index_eq(eq_name)
            index_io = self.get_index_io(io_name)
            index = np.intersect1d(index_io, index_eq)
            if conf_id is not None:
                index_conf = self.get_index_conf(conf_id)
                index = np.intersect1d(index, index_conf)
                if len(index) != 1:
                    raise ValueError('{} match found instead of 1 for EQ name {}, I/O name {} and configuration id {}'
                                     .format(len(index), eq_name, io_name, conf_id))

            if len(index) != 1:
                raise ValueError('{} match found instead of 1 for EQ name {} and I/O name {}'
                                 .format(len(index), eq_name, io_name))

            list(self.sheet.columns)[self.n_col][row + 1].value = index[0] + self.index_offset

    def get_index_eq(self, eq_name):
        return np.where(self.lua['interface_device_name'] == eq_name)[0]

    def get_index_io(self, io_name):
        return np.where(self.lua['full_name'] == io_name)[0]

    def get_index_conf(self, conf_id):
        return np.where(self.lua['standard_configuration_id'] == conf_id)[0]

    def get_difference(self, reference_key, compare, index, cross=''):
        """
        Makes the comparison between a cell from the LUA with a cell from the autocad file
        """
        str_compare = str(compare).replace('\P', '')
        # If there is no cross in the reference table
        if not isinstance(cross, str):
            return '', None

        str_reference = str(self.lua[reference_key][index])

        # If the LUA reference is empty
        if str_reference == 'nan':
            return 'Empty reference', self.orange

        # If the reference and the autocad output are identical
        if str_reference == str_compare:
            return 'ok', self.green

        # If there is a difference between the tables
        if str_compare == 'nan':
            return str_reference, self.yellow

        return str_reference, self.red

    def fill_cell(self, reference_key, compare, index, column, row, cross=''):
        content, color = self.get_difference(reference_key, compare, index, cross)
        list(self.sheet.columns)[column][row].value = content
        if color is not None:
            list(self.sheet.columns)[column][row].fill = color

    def compare(self):
        """
        Compares two matching rows of the LUA and autocad file
        """
        index_list = list(self.sheet.columns)[self.n_col]
        for row, (cell, dutch, french, signal, polarity) in \
            enumerate(zip(index_list[1:],
                          self.autocad['i/o_descr_nl'],
                          self.autocad['i/o_descr_fr'],
                          self.autocad['i/o_conn_02_function'],
                          self.autocad['i/o_conn_01_function'])):
            if row > self.last_row:
                break

            index = int(cell.value) - self.index_offset
            cross = self.get_cross(index)
            self.fill_cell('infolist/princ.diagr._name_nl', dutch, index, self.n_col + 1, row + 1)
            self.fill_cell('infolist/princ.diagr._name_fr', french, index, self.n_col + 2, row + 1)
            self.fill_cell('signal_code', signal, index, self.n_col + 3, row + 1, cross)
            self.fill_cell('polarity', polarity, index, self.n_col + 4, row + 1, cross)

    def get_cross(self, index):
        return self.lua[self.box_name][index]

    def adjust_columns(self):
        worksheet = self.results.active
        for n, col in enumerate(worksheet.columns):
            max_size = 0
            for cell in col:
                len_cell = len(str(cell.value))
                if len_cell > max_size:
                    max_size = len_cell

            column = get_column_letter(col[0].column)
            worksheet.column_dimensions[column].width = max_size + 5

    def save_result(self, output_path):
        try:
            self.results.save(output_path)
        except FileNotFoundError:
            raise ValueError('Path to results file {} not found'.format(output_path))
